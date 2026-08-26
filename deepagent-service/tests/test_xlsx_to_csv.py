from datetime import datetime
from pathlib import Path

import openpyxl

from app.engine.xlsx_to_csv import convert_xlsx_to_csv


def _write_xlsx(path: Path, rows: list[list[object]], extra_sheet: bool = False) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    if extra_sheet:
        second = workbook.create_sheet("second")
        second.append(["should", "not", "appear"])
    workbook.save(path)


def test_convert_first_sheet_values_to_csv(tmp_path: Path) -> None:
    source = tmp_path / "input.xlsx"
    _write_xlsx(source, [["name", "count"], ["widget", 3], ["gadget", 1.5]])
    output = tmp_path / "out.csv"
    convert_xlsx_to_csv(source, output)
    assert output.read_text(encoding="utf-8").splitlines() == [
        "name,count",
        "widget,3",
        "gadget,1.5",
    ]


def test_convert_only_first_sheet(tmp_path: Path) -> None:
    source = tmp_path / "multi.xlsx"
    _write_xlsx(source, [["a"]], extra_sheet=True)
    output = tmp_path / "out.csv"
    convert_xlsx_to_csv(source, output)
    assert "appear" not in output.read_text(encoding="utf-8")


def test_convert_none_and_datetime_cells(tmp_path: Path) -> None:
    source = tmp_path / "mixed.xlsx"
    _write_xlsx(source, [["when", "note"], [datetime(2026, 8, 26, 9, 30), None]])  # noqa: DTZ001
    output = tmp_path / "out.csv"
    convert_xlsx_to_csv(source, output)
    lines = output.read_text(encoding="utf-8").splitlines()
    assert lines[1] == "2026-08-26 09:30:00,"  # datetime→isoformat(sep=' ')、None→空字串


def test_convert_invalid_xlsx_raises(tmp_path: Path) -> None:
    source = tmp_path / "garbage.xlsx"
    source.write_bytes(b"not a zip at all")
    import pytest

    with pytest.raises(Exception):  # noqa: B017 fail loud:internal 未解密的密文走到這裡必炸
        convert_xlsx_to_csv(source, tmp_path / "out.csv")


def test_convert_stale_dimension_trims_phantom_rows_and_columns(tmp_path: Path) -> None:
    source = tmp_path / "stale_dimension.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["name", "count"])
    sheet.append(["widget", 3])
    sheet["F10"] = "x"
    sheet["F10"] = None
    workbook.save(source)

    # 先證明 fixture 真的觸發了失真的 dimension,測試才不會在「自癒」的 fixture 上悄悄通過。
    verify_workbook = openpyxl.load_workbook(source, read_only=True)
    verify_sheet = verify_workbook.active
    inflated_max_row = verify_sheet.max_row
    inflated_max_column = verify_sheet.max_column
    verify_workbook.close()
    assert (inflated_max_row, inflated_max_column) == (10, 6)

    output = tmp_path / "out.csv"
    convert_xlsx_to_csv(source, output)
    lines = output.read_text(encoding="utf-8").splitlines()
    assert lines == ["name,count", "widget,3"]
    assert all(len(line.split(",")) == 2 for line in lines)


def test_convert_boolean_cells_render_uppercase(tmp_path: Path) -> None:
    source = tmp_path / "bools.xlsx"
    _write_xlsx(source, [[True, False]])
    output = tmp_path / "out.csv"
    convert_xlsx_to_csv(source, output)
    assert output.read_text(encoding="utf-8").splitlines() == ["TRUE,FALSE"]


def test_convert_empty_sheet_writes_empty_csv(tmp_path: Path) -> None:
    source = tmp_path / "empty.xlsx"
    workbook = openpyxl.Workbook()
    workbook.save(source)  # 預設空白 sheet,不 append 任何 row
    output = tmp_path / "out.csv"
    convert_xlsx_to_csv(source, output)
    assert output.read_text(encoding="utf-8") == ""
