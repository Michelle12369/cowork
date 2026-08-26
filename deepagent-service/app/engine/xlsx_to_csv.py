"""xlsx→CSV 轉檔(openpyxl read-only streaming)。

語意對齊 Java 舊管線(POI DataFormatter):僅第一張 sheet、cell 輸出為文字。
已知差異(允收):數字格式代碼(千分位/百分比樣式)不重現——輸出原始值文字;
datetime 輸出 `YYYY-MM-DD HH:MM:SS`;布林輸出 `TRUE`/`FALSE`(對齊 POI)。

worksheet 快取的 `<dimension>` 範圍可能因「先寫入格式/值後又清空」而失真,
留下全空的幻影列/欄——若照單全收會讓下游 schema 被幻影欄污染。故採兩遍
streaming 掃描(read-only worksheet 支援重複 iterate,對大檔仍是常數記憶體):
第一遍只找出真正有資料的最後一列與最大欄,第二遍依此範圍截斷/補齊輸出。

壞檔(未解密密文/非 xlsx)由 openpyxl 直接 raise——fail loud 是設計要求,
勿包成靜默略過。
"""

import csv
from datetime import datetime
from pathlib import Path

import openpyxl


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        # MUST 排在數字判斷之前:Python 的 bool 是 int 子類。
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def convert_xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> None:
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        first_sheet = workbook.worksheets[0]

        last_data_row = 0
        max_data_column = 0
        for row_index, row in enumerate(first_sheet.iter_rows(values_only=True), start=1):
            for column_index, cell in enumerate(row, start=1):
                if cell is not None:
                    last_data_row = row_index
                    max_data_column = max(max_data_column, column_index)

        with csv_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            if last_data_row == 0:
                return
            for row_index, row in enumerate(first_sheet.iter_rows(values_only=True), start=1):
                if row_index > last_data_row:
                    break
                trimmed_row = list(row[:max_data_column])
                trimmed_row += [None] * (max_data_column - len(trimmed_row))
                writer.writerow([_format_cell(cell) for cell in trimmed_row])
    finally:
        workbook.close()
